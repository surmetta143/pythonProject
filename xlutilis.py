import openpyxl
class ReadData:
  def getRowCount(file):
    work_book = openpyxl.load_workbook(file)
    sheet= work_book.active
    return sheet.max_row
  def getColCount(file):
    work_book =openpyxl.load_workbook(file)
    sheet= work_book.active
    return sheet.max_column
  def ReadData(file,r,c):
    work_book = openpyxl.load_workbook(file)
    sheet = work_book.active
    return sheet.cell(row=r,column=c).value

  def WriteData(file,r,c,data):
    work_book= openpyxl.load_workbook(file)
    sheet = work_book.active
    sheet.cell(row=r,column=c).value=data
    work_book.save(file)
